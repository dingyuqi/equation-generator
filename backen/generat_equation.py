import os.path
import random
import string
import time

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

MAX_ROW = 10


def generate_equations(num_equations, operators, min_num, max_num, num_terms=2):
    """
    This function generates a specified number of arithmetic equations with random operands and operators.

    Parameters:
        - num_equations (int): the number of equations to generate
        - min_num (int): the minimum value for the operands
        - max_num (int): the maximum value for the operands
        - num_terms (int): the number of terms in each equation (default is 2)

    Returns:
        - equations (list): a list of randomly generated equations
    """
    equations = []

    for i in range(num_equations):
        equation = str(random.randint(min_num, max_num))

        for j in range(num_terms - 1):
            operator = random.choice(operators)
            operand = str(random.randint(min_num, max_num))
            equation += " {} {}".format(operator, operand)
            if j == num_terms - 2:
                equation += "="

        equations.append(equation)

    return equations


def write_excel(dir_path, data):
    timestamp = time.time()
    file_path = dir_path + '/' + "算式" + str(int(timestamp)) + ".xlsx"
    wb = Workbook()
    # 先不指定，最后再保存
    ws = wb.active

    # 指定位置为第一行中的不同的列，值为value
    row = 1
    col = 1
    count = 0
    for ele in data:
        if count == MAX_ROW:
            count = 0
            col += 1
            row = 1
        ws.cell(row=row, column=col, value=ele)
        row += 1
        count += 1

    wb.save(file_path)
